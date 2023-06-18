#! python3
# 機能
#  テキストファイルをExcelに変換する
# 使い方
#  1.Pythonを実行する
# 実行コマンド
#  python txt2xlsx.py 入力フォルダ 出力ファイル名
#  python txt2xlsx.py input output.xlsx

import os
import sys
import openpyxl

# テキストファイルをExcelに変換する関数
def txt2xlsx(input):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # フォルダ内のテキストファイルを探す
    col = 0


    with open("./" + input, "r", encoding="utf-8") as text_file:
        content = text_file.read()
        # for line in text_file:
        #     if line[0] == '+':
        #         continue
        #     print(line)
    print(type(content))
    content = content.replace('+','').replace('-','')
    content = content.split('|')
    row = 1
    col = 1
    for ch in content:
        if ch.strip() == '':
            print("none")
            row += 1
            col = 1
            continue
        print(ch.strip())
        sheet.cell(column=col, row=row).value = ch.strip()
        # シートの列ごとにテキストファイルの内容を書き込む
        col += 1
        # for row, text in enumerate(text_file):
        #     sheet.cell(column=col, row=row + 1).value = text.strip()
    print(input.rstrip('.txt'))
    wb.save(input.rstrip('.txt') + ".xlsx")
args = sys.argv
txt2xlsx(args[1])